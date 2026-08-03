# -*- coding: utf-8 -*-
"""
屏幕区域 OCR 自动抓取工具
============================================
流程: 截取屏幕指定区域 -> Umi-OCR 识别文字 -> 写入 Excel -> 点击"下一页"按钮 -> 循环 N 次

使用前提:
  - 程序(exe)旁边需要有一个包含 Umi-OCR.exe 的文件夹(官方发行版解压后即为 Umi-OCR 文件夹)。
    程序启动时会自动找到并拉起它, 无需手动打开。
  - Umi-OCR 的本地 HTTP 服务接口 (http://127.0.0.1:1224) 默认开启。

构建 exe(在任意 Windows 电脑上, 需已装 Python 3.8):
  pip install pyautogui pillow openpyxl requests pyinstaller==5.13.2
  pyinstaller --noconfirm --clean --onedir --windowed --name ScreenOCR app/main.py
  或直接双击 build/build_win.bat

安全停止: 运行过程中把鼠标快速甩到屏幕左上角(0,0)附近, 程序立即中止。
"""
import base64
import io
import os
import subprocess
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox

import pyautogui
import requests
from openpyxl import Workbook, load_workbook

APP_TITLE = "屏幕OCR自动抓取工具"

# ------------------------- 可调参数(一般不需要改) -------------------------
UmiOCR_HOST = "http://127.0.0.1"
UmiOCR_PORT = 1224
OCR_READY_TIMEOUT = 90          # 等待 OCR 引擎启动的最长时间(秒)
OCR_REQUEST_TIMEOUT = 120       # 单次 OCR 请求超时(秒)
OCR_RETRY = 3                   # OCR 请求失败重试次数
PAGE_WAIT = 2.0                 # 点击"下一页"后等待页面刷新的秒数
OCR_OPTIONS = {                 # Umi-OCR 识别参数
    "ocr.language": "models/config_chinese.txt",  # 简体中文模型
    "data.format": "text",                        # 只返回文本
    "tbpu.parser": "single_line",                 # 单栏-总是换行: 每行文字单独一行
    "ocr.cls": True,                              # 纠正文本方向
}
EXCEL_HEADER = ["页码", "行号", "识别内容"]
# ---------------------------------------------------------------------------


def app_dir():
    """程序所在目录(exe 运行时为其所在目录)。"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def find_umi_exe():
    """在程序旁边查找 Umi-OCR.exe(支持多层子目录)。"""
    base = app_dir()
    found = []
    for root, dirs, files in os.walk(base):
        if "Umi-OCR.exe" in files:
            found.append(os.path.join(root, "Umi-OCR.exe"))
        depth = root[len(base):].count(os.sep)
        if depth >= 2:
            dirs[:] = []
    if found:
        return sorted(found)[0]
    return None


class UmiEngine:
    """Umi-OCR 本地引擎: 自动启动 + HTTP 调用。"""

    def __init__(self, log):
        self.log = log
        self.proc = None
        self.path = None

    def url(self, api):
        return "%s:%d%s" % (UmiOCR_HOST, UmiOCR_PORT, api)

    def is_up(self):
        try:
            requests.get(self.url("/api/ocr/get_options"), timeout=2)
            return True
        except Exception:
            return False

    def start(self):
        """确保引擎运行, 返回 True/False。"""
        if self.is_up():
            self.log("OCR 引擎已在运行(端口 %d), 直接使用" % UmiOCR_PORT)
            return True
        self.path = find_umi_exe()
        if not self.path:
            return False
        self.log("正在启动 OCR 引擎: %s" % self.path)
        self.proc = subprocess.Popen([self.path], cwd=os.path.dirname(self.path))
        deadline = time.time() + OCR_READY_TIMEOUT
        while time.time() < deadline:
            if self.is_up():
                self.log("OCR 引擎就绪")
                return True
            time.sleep(0.5)
        return False

    def stop(self):
        if self.proc and self.proc.poll() is None:
            try:
                self.proc.terminate()
            except Exception:
                pass
        self.proc = None

    def ocr(self, img):
        """识别一张图片, 返回识别文本(每行以\\n分隔)。识别失败抛异常。"""
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        payload = {"base64": b64, "options": OCR_OPTIONS}
        last = ""
        for i in range(OCR_RETRY):
            try:
                r = requests.post(self.url("/api/ocr"), json=payload,
                                  timeout=OCR_REQUEST_TIMEOUT)
                res = r.json()
                code = res.get("code")
                if code == 100:
                    return res.get("data", "")
                if code == 101:
                    return ""
                last = "引擎返回错误 code=%s: %s" % (code, res.get("data"))
            except Exception as e:
                last = "请求异常: %s" % e
            if i < OCR_RETRY - 1:
                time.sleep(1.5)
        raise RuntimeError("OCR 识别失败: %s" % last)


class Overlay:
    """全屏遮罩: 拖拽框选截图区域 / 单击定位点击位置。"""

    def __init__(self, app, hint, mode, callback):
        self.app = app
        self.callback = callback
        self.mode = mode          # "region" 或 "click"
        self.top = tk.Toplevel(app.root)
        self.top.overrideredirect(True)
        self.top.attributes("-topmost", True)
        try:
            self.top.attributes("-alpha", 0.30)
        except Exception:
            pass
        w = app.root.winfo_screenwidth()
        h = app.root.winfo_screenheight()
        self.top.geometry("%dx%d+0+0" % (w, h))
        self.cv = tk.Canvas(self.top, bg="black", cursor="crosshair",
                            highlightthickness=0)
        self.cv.pack(fill="both", expand=True)
        self.cv.create_text(w // 2, 26, text=hint, fill="white",
                            font=("Microsoft YaHei", 15, "bold"))
        self.info = self.cv.create_text(8, h - 14, anchor="sw", text="",
                                        fill="#00ff90",
                                        font=("Microsoft YaHei", 12))
        self.start_xy = None
        self.rect_id = None
        self.cv.bind("<ButtonPress-1>", self._press)
        self.cv.bind("<B1-Motion>", self._drag)
        self.cv.bind("<ButtonRelease-1>", self._release)
        self.cv.bind("<Escape>", lambda e: self._finish(None))
        self.top.focus_force()
        self.top.grab_set()
        self._update_info(0, 0)

    def _update_info(self, x, y):
        self.cv.itemconfig(self.info, text="鼠标位置: %d, %d   Esc 取消" % (x, y))

    def _press(self, e):
        self.start_xy = (e.x_root, e.y_root)
        if self.mode == "click":
            self.cv.create_oval(e.x_root - 5, e.y_root - 5,
                                e.x_root + 5, e.y_root + 5,
                                fill="#ff4040", outline="")
            self._update_info(e.x_root, e.y_root)

    def _drag(self, e):
        self._update_info(e.x_root, e.y_root)
        if self.mode == "region" and self.start_xy:
            if self.rect_id:
                self.cv.delete(self.rect_id)
            x1, y1 = self.start_xy
            self.rect_id = self.cv.create_rectangle(
                x1, y1, e.x_root, e.y_root,
                outline="#ff4040", width=2, fill="#4040ff", stipple="gray50")

    def _release(self, e):
        if self.mode == "region":
            if not self.start_xy:
                self._finish(None)
                return
            x1, y1 = self.start_xy
            x2, y2 = e.x_root, e.y_root
            l, t = min(x1, x2), min(y1, y2)
            w, h = abs(x2 - x1), abs(y2 - y1)
            if w < 3 or h < 3:
                self._finish(None)
                return
            self._finish((l, t, w, h))
        else:
            self._finish((e.x_root, e.y_root))

    def _finish(self, result):
        try:
            self.top.grab_release()
        except Exception:
            pass
        self.top.destroy()
        if result is not None:
            self.callback(result)


class App:
    def __init__(self, root):
        self.root = root
        self.engine = UmiEngine(self.log)
        self.region = None        # (left, top, width, height)
        self.click_pos = None     # (x, y)
        self.excel_path = None
        self.stop_flag = threading.Event()
        self.running = False
        root.title(APP_TITLE)
        root.resizable(False, False)

        pad = {"padx": 10, "pady": 4}
        frm = tk.Frame(root)
        frm.pack(padx=14, pady=10)

        tk.Label(frm, text="循环次数(要抓取多少页):").grid(
            row=0, column=0, sticky="w", **pad)
        self.var_times = tk.StringVar(value="10")
        tk.Entry(frm, textvariable=self.var_times, width=10).grid(
            row=0, column=1, sticky="w", **pad)
        tk.Label(frm, text="(运行中把鼠标甩到屏幕左上角可紧急停止)").grid(
            row=0, column=2, sticky="w", **pad)

        self.btn_region = tk.Button(frm, text="① 选择截图区域", width=18,
                                    command=self.select_region)
        self.btn_region.grid(row=1, column=0, sticky="w", **pad)
        self.lbl_region = tk.Label(frm, text="未设置", fg="gray")
        self.lbl_region.grid(row=1, column=1, columnspan=2, sticky="w", **pad)

        self.btn_click = tk.Button(frm, text="② 定位下一页按钮", width=18,
                                   command=self.select_click)
        self.btn_click.grid(row=2, column=0, sticky="w", **pad)
        self.lbl_click = tk.Label(frm, text="未设置", fg="gray")
        self.lbl_click.grid(row=2, column=1, columnspan=2, sticky="w", **pad)

        self.btn_excel = tk.Button(frm, text="③ 选择Excel保存文件", width=18,
                                   command=self.select_excel)
        self.btn_excel.grid(row=3, column=0, sticky="w", **pad)
        self.lbl_excel = tk.Label(frm, text="未设置", fg="gray")
        self.lbl_excel.grid(row=3, column=1, columnspan=2, sticky="w", **pad)

        btns = tk.Frame(frm)
        btns.grid(row=4, column=0, columnspan=3, pady=8)
        self.btn_start = tk.Button(btns, text="▶ 开始抓取", width=14,
                                   bg="#e8ffe8", command=self.start)
        self.btn_start.pack(side="left", padx=6)
        self.btn_stop = tk.Button(btns, text="■ 停止", width=10,
                                  state="disabled", command=self.stop)
        self.btn_stop.pack(side="left", padx=6)

        tk.Label(frm, text="运行日志:").grid(row=5, column=0, sticky="w", **pad)
        self.txt = tk.Text(frm, width=72, height=12, state="disabled",
                           font=("Courier New", 10))
        self.txt.grid(row=6, column=0, columnspan=3, **pad)

        root.protocol("WM_DELETE_WINDOW", self.on_close)

    # ---------------- GUI 动作 ----------------
    def log(self, msg):
        try:
            self.root.after(0, self._log_now, msg)
        except Exception:
            pass

    def _log_now(self, msg):
        self.txt.config(state="normal")
        self.txt.insert("end", time.strftime("[%H:%M:%S] ") + msg + "\n")
        self.txt.see("end")
        self.txt.config(state="disabled")

    def select_region(self):
        if self.running:
            return
        Overlay(self, "按住鼠标左键拖动, 框选要识别的屏幕区域, 松开完成; Esc 取消",
                "region", self._on_region)

    def _on_region(self, region):
        self.region = region
        self.lbl_region.config(text="(%d, %d)  宽%d 高%d" % region,
                               fg="black")
        self.log("截图区域: " + str(region))

    def select_click(self):
        if self.running:
            return
        Overlay(self, "移动鼠标到\"下一页\"按钮上, 单击左键记录位置; Esc 取消",
                "click", self._on_click)

    def _on_click(self, pos):
        self.click_pos = pos
        self.lbl_click.config(text="(%d, %d)" % pos, fg="black")
        self.log("点击位置: " + str(pos))

    def select_excel(self):
        if self.running:
            return
        path = filedialog.asksaveasfilename(
            title="选择 Excel 保存文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="识别结果.xlsx")
        if not path:
            return
        self.excel_path = path
        self.lbl_excel.config(text=path, fg="black")
        self.log("Excel 文件: " + path)

    # ---------------- 开始 / 停止 ----------------
    def start(self):
        try:
            times = int(self.var_times.get())
        except ValueError:
            messagebox.showerror(APP_TITLE, "循环次数必须是数字")
            return
        if times < 1:
            messagebox.showerror(APP_TITLE, "循环次数至少为 1")
            return
        if not self.region:
            messagebox.showerror(APP_TITLE, "请先选择截图区域")
            return
        if not self.click_pos:
            messagebox.showerror(APP_TITLE, "请先定位\"下一页\"按钮")
            return
        if not self.excel_path:
            messagebox.showerror(APP_TITLE, "请先选择 Excel 保存文件")
            return
        self.running = True
        self.stop_flag.clear()
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.var_times.get()
        threading.Thread(target=self._run, args=(times,), daemon=True).start()

    def stop(self):
        self.stop_flag.set()
        self.log("收到停止指令, 正在结束…")

    def on_close(self):
        self.stop_flag.set()
        self.engine.stop()
        self.root.destroy()

    # ---------------- 主流程 ----------------
    def _run(self, times):
        try:
            pyautogui.FAILSAFE = True
            ok = self.engine.start()
            if not ok:
                self.log("!! 未找到/无法启动 OCR 引擎")
                self.log("  请确认程序文件夹内包含 Umi-OCR 文件夹(内含 Umi-OCR.exe)")
                self.log("  也可以手动打开 Umi-OCR 后, 重新点击开始")
                self._done(False)
                return
            wb, ws = self._open_excel()
            done = 0
            for i in range(1, times + 1):
                if self.stop_flag.is_set():
                    break
                try:
                    self.log("第 %d/%d 页: 截图…" % (i, times))
                    img = pyautogui.screenshot(region=self.region)
                    text = self.engine.ocr(img)
                except Exception as e:
                    self.log("!! 第 %d 页失败: %s" % (i, e))
                    break
                lines = [ln for ln in text.split("\n")] if text else [""]
                for j, line in enumerate(lines, 1):
                    ws.append([i, j, line])
                wb.save(self.excel_path)
                self.log("第 %d 页: 识别 %d 行, 已写入 Excel" % (i, len(lines)))
                done = i
                if i < times and not self.stop_flag.is_set():
                    self.log("点击\"下一页\"…")
                    pyautogui.click(*self.click_pos)
                    time.sleep(PAGE_WAIT)
            if done:
                self.log("全部完成, 共写入 %d 页到:\n    %s" % (done, self.excel_path))
            self._done(True)
        except pyautogui.FailSafeException:
            self.log("!! 已触发紧急停止(鼠标甩到左上角)")
            self._done(False)
        except Exception as e:
            self.log("!! 运行出错: %s" % e)
            self._done(False)

    def _open_excel(self):
        """打开或新建 Excel, 返回 (wb, ws)。已有文件则追加。"""
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            ws = wb.active
            if not ws.max_row or ws["A1"].value != EXCEL_HEADER[0]:
                ws.insert_rows(1)
                ws.append(EXCEL_HEADER)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "识别结果"
            ws.append(EXCEL_HEADER)
        return wb, ws

    def _done(self, normal):
        self.engine.stop()
        self.running = False
        self.root.after(0, self._done_now, normal)

    def _done_now(self, normal):
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        if normal:
            messagebox.showinfo(APP_TITLE, "任务完成, 结果已保存到 Excel 文件")
        else:
            messagebox.showwarning(APP_TITLE, "任务已中止, 请查看运行日志")


def main():
    root = tk.Tk()
    root.option_add("*font", ("Microsoft YaHei", 10))
    App(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        try:
            with open(os.path.join(app_dir(), "run.log"), "w",
                      encoding="utf-8") as f:
                f.write("%s\n%s" % (e, sys.exc_info()[2]))
        except Exception:
            pass
        try:
            messagebox.showerror(APP_TITLE, "程序出错:\n%s" % e)
        except Exception:
            pass
